"""Trade journal with P&L tracking."""
from __future__ import annotations
import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional
from dataclasses import dataclass, asdict, field
from enum import Enum

from .paths import user_data_dir

_log = logging.getLogger(__name__)


class TradeStatus(Enum):
    OPEN = "open"
    CLOSED = "closed"
    CANCELLED = "cancelled"
    PARTIAL = "partial"


class TradeSide(Enum):
    LONG = "long"
    SHORT = "short"


@dataclass
class TradeEntry:
    id: str
    symbol: str
    side: str  # "long" or "short"
    status: str  # "open", "closed", "cancelled", "partial"

    # Entry details
    entry_time: str
    entry_price: float
    quantity: int

    # Exit details (filled when closed)
    exit_time: Optional[str] = None
    exit_price: Optional[float] = None
    exit_quantity: Optional[int] = None

    # Order IDs
    parent_order_id: Optional[int] = None
    stop_order_id: Optional[int] = None
    take_order_id: Optional[int] = None

    # Levels
    stop_price: Optional[float] = None
    take_profit_price: Optional[float] = None

    # P&L
    realized_pnl: Optional[float] = None
    commission: float = 0.0

    # Metadata
    strategy: str = "swing_pullback_atr"
    notes: str = ""
    tags: List[str] = field(default_factory=list)

    # Plan reference
    plan_file: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return asdict(self)

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'TradeEntry':
        """Create from dictionary."""
        # Handle missing fields for backwards compatibility
        if 'tags' not in data:
            data['tags'] = []
        return cls(**data)

    @property
    def is_open(self) -> bool:
        return self.status == TradeStatus.OPEN.value

    @property
    def is_closed(self) -> bool:
        return self.status == TradeStatus.CLOSED.value

    @property
    def risk_per_share(self) -> Optional[float]:
        if self.stop_price and self.entry_price:
            return abs(self.entry_price - self.stop_price)
        return None

    @property
    def r_multiple(self) -> Optional[float]:
        """Calculate R-multiple if trade is closed."""
        if not self.is_closed or not self.realized_pnl:
            return None
        risk = self.risk_per_share
        if risk and risk > 0:
            return self.realized_pnl / (risk * self.quantity)
        return None


class TradeJournal:
    def __init__(self, journal_dir: Optional[Path] = None):
        if journal_dir is None:
            journal_dir = user_data_dir() / "journal"
        self._journal_dir = journal_dir
        self._journal_dir.mkdir(parents=True, exist_ok=True)
        self._journal_file = self._journal_dir / "trades.json"
        self._trades: Dict[str, TradeEntry] = {}
        self._load()

    def _load(self) -> None:
        """Load trades from disk."""
        if self._journal_file.exists():
            try:
                with open(self._journal_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                for trade_id, trade_data in data.get('trades', {}).items():
                    self._trades[trade_id] = TradeEntry.from_dict(trade_data)
                _log.info(f"Loaded {len(self._trades)} trades from journal")
            except Exception as e:
                _log.error(f"Error loading trade journal: {e}")
                self._trades = {}
        else:
            _log.debug("No existing trade journal found")

    def _save(self) -> None:
        """Save trades to disk with automatic backup."""
        try:
            # Create backup before overwriting
            if self._journal_file.exists():
                self._create_backup()

            data = {
                'version': '1.0',
                'updated_at': datetime.now(timezone.utc).isoformat(),
                'trades': {tid: t.to_dict() for tid, t in self._trades.items()}
            }
            with open(self._journal_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)
            _log.debug(f"Saved {len(self._trades)} trades to journal")
        except Exception as e:
            _log.error(f"Error saving trade journal: {e}")

    def _create_backup(self) -> None:
        """Create a rolling backup of the journal."""
        try:
            backup_dir = self._journal_dir / "backups"
            backup_dir.mkdir(parents=True, exist_ok=True)

            # Keep last 5 backups
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = backup_dir / f"trades_{timestamp}.json"

            import shutil
            shutil.copy2(self._journal_file, backup_file)

            # Clean old backups (keep last 5)
            backups = sorted(backup_dir.glob("trades_*.json"), reverse=True)
            for old_backup in backups[5:]:
                old_backup.unlink()

            _log.debug(f"Created journal backup: {backup_file.name}")
        except Exception as e:
            _log.warning(f"Failed to create journal backup: {e}")

    def _generate_id(self) -> str:
        """Generate a unique trade ID."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        count = len(self._trades) + 1
        return f"T{timestamp}_{count:04d}"

    def add_trade(
        self,
        symbol: str,
        side: str,
        entry_price: float,
        quantity: int,
        stop_price: Optional[float] = None,
        take_profit_price: Optional[float] = None,
        parent_order_id: Optional[int] = None,
        stop_order_id: Optional[int] = None,
        take_order_id: Optional[int] = None,
        strategy: str = "swing_pullback_atr",
        notes: str = "",
        plan_file: Optional[str] = None,
    ) -> TradeEntry:
        """Add a new trade to the journal."""
        trade_id = self._generate_id()
        trade = TradeEntry(
            id=trade_id,
            symbol=symbol,
            side=side,
            status=TradeStatus.OPEN.value,
            entry_time=datetime.now(timezone.utc).isoformat(),
            entry_price=entry_price,
            quantity=quantity,
            stop_price=stop_price,
            take_profit_price=take_profit_price,
            parent_order_id=parent_order_id,
            stop_order_id=stop_order_id,
            take_order_id=take_order_id,
            strategy=strategy,
            notes=notes,
            plan_file=plan_file,
        )
        self._trades[trade_id] = trade
        self._save()
        _log.info(f"Added trade {trade_id}: {side} {quantity} {symbol} @ {entry_price}")
        return trade

    def close_trade(
        self,
        trade_id: str,
        exit_price: float,
        exit_quantity: Optional[int] = None,
        realized_pnl: Optional[float] = None,
        commission: float = 0.0,
    ) -> Optional[TradeEntry]:
        """Close an existing trade."""
        trade = self._trades.get(trade_id)
        if not trade:
            _log.warning(f"Trade {trade_id} not found")
            return None

        trade.exit_time = datetime.now(timezone.utc).isoformat()
        trade.exit_price = exit_price
        trade.exit_quantity = exit_quantity or trade.quantity
        trade.commission = commission

        # Calculate P&L if not provided
        if realized_pnl is not None:
            trade.realized_pnl = realized_pnl
        else:
            if trade.side == TradeSide.LONG.value:
                trade.realized_pnl = (exit_price - trade.entry_price) * trade.exit_quantity - commission
            else:
                trade.realized_pnl = (trade.entry_price - exit_price) * trade.exit_quantity - commission

        # Update status
        if trade.exit_quantity >= trade.quantity:
            trade.status = TradeStatus.CLOSED.value
        else:
            trade.status = TradeStatus.PARTIAL.value

        self._save()
        _log.info(f"Closed trade {trade_id}: P&L ${trade.realized_pnl:.2f}")
        return trade

    def cancel_trade(self, trade_id: str, reason: str = "") -> Optional[TradeEntry]:
        """Mark a trade as cancelled."""
        trade = self._trades.get(trade_id)
        if not trade:
            return None

        trade.status = TradeStatus.CANCELLED.value
        if reason:
            trade.notes = f"{trade.notes}\nCancelled: {reason}".strip()
        self._save()
        _log.info(f"Cancelled trade {trade_id}")
        return trade

    def update_notes(self, trade_id: str, notes: str) -> Optional[TradeEntry]:
        """Update trade notes."""
        trade = self._trades.get(trade_id)
        if trade:
            trade.notes = notes
            self._save()
        return trade

    def add_tag(self, trade_id: str, tag: str) -> Optional[TradeEntry]:
        """Add a tag to a trade."""
        trade = self._trades.get(trade_id)
        if trade and tag not in trade.tags:
            trade.tags.append(tag)
            self._save()
        return trade

    def get_trade(self, trade_id: str) -> Optional[TradeEntry]:
        """Get a trade by ID."""
        return self._trades.get(trade_id)

    def get_trade_by_order_id(self, order_id: int) -> Optional[TradeEntry]:
        """Find a trade by any of its order IDs."""
        for trade in self._trades.values():
            if order_id in (trade.parent_order_id, trade.stop_order_id, trade.take_order_id):
                return trade
        return None

    def get_all_trades(self) -> List[TradeEntry]:
        """Get all trades, sorted by entry time (newest first)."""
        return sorted(self._trades.values(), key=lambda t: t.entry_time, reverse=True)

    def get_open_trades(self) -> List[TradeEntry]:
        """Get all open trades."""
        return [t for t in self._trades.values() if t.is_open]

    def get_closed_trades(self) -> List[TradeEntry]:
        """Get all closed trades."""
        return [t for t in self._trades.values() if t.is_closed]

    def get_trades_by_symbol(self, symbol: str) -> List[TradeEntry]:
        """Get all trades for a specific symbol."""
        return [t for t in self._trades.values() if t.symbol == symbol]

    def get_statistics(self) -> Dict[str, Any]:
        """Calculate trading statistics."""
        closed = self.get_closed_trades()
        if not closed:
            return {
                'total_trades': len(self._trades),
                'open_trades': len(self.get_open_trades()),
                'closed_trades': 0,
                'win_rate': 0.0,
                'total_pnl': 0.0,
                'avg_pnl': 0.0,
                'avg_winner': 0.0,
                'avg_loser': 0.0,
                'best_trade': 0.0,
                'worst_trade': 0.0,
                'profit_factor': 0.0,
                'avg_r_multiple': 0.0,
            }

        winners = [t for t in closed if t.realized_pnl and t.realized_pnl > 0]
        losers = [t for t in closed if t.realized_pnl and t.realized_pnl < 0]

        total_pnl = sum(t.realized_pnl or 0 for t in closed)
        gross_profit = sum(t.realized_pnl for t in winners) if winners else 0
        gross_loss = abs(sum(t.realized_pnl for t in losers)) if losers else 0

        r_multiples = [t.r_multiple for t in closed if t.r_multiple is not None]

        return {
            'total_trades': len(self._trades),
            'open_trades': len(self.get_open_trades()),
            'closed_trades': len(closed),
            'winners': len(winners),
            'losers': len(losers),
            'win_rate': len(winners) / len(closed) * 100 if closed else 0,
            'total_pnl': total_pnl,
            'avg_pnl': total_pnl / len(closed) if closed else 0,
            'avg_winner': gross_profit / len(winners) if winners else 0,
            'avg_loser': -gross_loss / len(losers) if losers else 0,
            'best_trade': max((t.realized_pnl or 0) for t in closed) if closed else 0,
            'worst_trade': min((t.realized_pnl or 0) for t in closed) if closed else 0,
            'profit_factor': gross_profit / gross_loss if gross_loss > 0 else float('inf'),
            'avg_r_multiple': sum(r_multiples) / len(r_multiples) if r_multiples else 0,
        }

    def export_to_csv(self, filepath: Path) -> bool:
        """Export trades to CSV file."""
        try:
            import csv
            trades = self.get_all_trades()
            if not trades:
                return False

            fieldnames = [
                'id', 'symbol', 'side', 'status', 'entry_time', 'entry_price',
                'quantity', 'exit_time', 'exit_price', 'stop_price',
                'take_profit_price', 'realized_pnl', 'commission', 'strategy', 'notes'
            ]

            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
                writer.writeheader()
                for trade in trades:
                    writer.writerow(trade.to_dict())

            _log.info(f"Exported {len(trades)} trades to {filepath}")
            return True
        except Exception as e:
            _log.error(f"Error exporting trades: {e}")
            return False

    def export_to_excel(self, filepath: Path) -> bool:
        """Export trades to Excel file with formatting."""
        try:
            trades = self.get_all_trades()
            if not trades:
                return False

            # Try openpyxl first, fall back to CSV if not available
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                _log.warning("openpyxl not installed, falling back to CSV export")
                csv_path = filepath.with_suffix('.csv')
                return self.export_to_csv(csv_path)

            wb = Workbook()
            ws = wb.active
            ws.title = "Trades"

            # Headers
            headers = [
                'ID', 'Symbol', 'Side', 'Status', 'Entry Time', 'Entry Price',
                'Qty', 'Exit Time', 'Exit Price', 'Stop', 'Take Profit',
                'P&L', 'Commission', 'R-Multiple', 'Strategy', 'Notes'
            ]

            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Data rows
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            for row_idx, trade in enumerate(trades, 2):
                ws.cell(row=row_idx, column=1, value=trade.id)
                ws.cell(row=row_idx, column=2, value=trade.symbol)
                ws.cell(row=row_idx, column=3, value=trade.side)
                ws.cell(row=row_idx, column=4, value=trade.status)
                ws.cell(row=row_idx, column=5, value=trade.entry_time)
                ws.cell(row=row_idx, column=6, value=trade.entry_price)
                ws.cell(row=row_idx, column=7, value=trade.quantity)
                ws.cell(row=row_idx, column=8, value=trade.exit_time or "")
                ws.cell(row=row_idx, column=9, value=trade.exit_price or "")
                ws.cell(row=row_idx, column=10, value=trade.stop_price or "")
                ws.cell(row=row_idx, column=11, value=trade.take_profit_price or "")

                pnl_cell = ws.cell(row=row_idx, column=12, value=trade.realized_pnl or "")
                if trade.realized_pnl:
                    if trade.realized_pnl > 0:
                        pnl_cell.fill = green_fill
                    elif trade.realized_pnl < 0:
                        pnl_cell.fill = red_fill

                ws.cell(row=row_idx, column=13, value=trade.commission)
                ws.cell(row=row_idx, column=14, value=round(trade.r_multiple, 2) if trade.r_multiple else "")
                ws.cell(row=row_idx, column=15, value=trade.strategy)
                ws.cell(row=row_idx, column=16, value=trade.notes)

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)

            # Add statistics sheet
            stats = self.get_statistics()
            ws_stats = wb.create_sheet("Statistics")
            stat_rows = [
                ("Total Trades", stats.get('total_trades', 0)),
                ("Open Trades", stats.get('open_trades', 0)),
                ("Closed Trades", stats.get('closed_trades', 0)),
                ("Winners", stats.get('winners', 0)),
                ("Losers", stats.get('losers', 0)),
                ("Win Rate", f"{stats.get('win_rate', 0):.1f}%"),
                ("Total P&L", f"${stats.get('total_pnl', 0):.2f}"),
                ("Avg P&L", f"${stats.get('avg_pnl', 0):.2f}"),
                ("Avg Winner", f"${stats.get('avg_winner', 0):.2f}"),
                ("Avg Loser", f"${stats.get('avg_loser', 0):.2f}"),
                ("Best Trade", f"${stats.get('best_trade', 0):.2f}"),
                ("Worst Trade", f"${stats.get('worst_trade', 0):.2f}"),
                ("Profit Factor", f"{stats.get('profit_factor', 0):.2f}"),
                ("Avg R-Multiple", f"{stats.get('avg_r_multiple', 0):.2f}R"),
            ]

            for row_idx, (label, value) in enumerate(stat_rows, 1):
                ws_stats.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                ws_stats.cell(row=row_idx, column=2, value=value)

            wb.save(filepath)
            _log.info(f"Exported {len(trades)} trades to {filepath}")
            return True
        except Exception as e:
            _log.error(f"Error exporting to Excel: {e}")
            return False


# Global instance
_journal: Optional[TradeJournal] = None


def get_trade_journal() -> TradeJournal:
    """Get the global trade journal instance."""
    global _journal
    if _journal is None:
        _journal = TradeJournal()
    return _journal
